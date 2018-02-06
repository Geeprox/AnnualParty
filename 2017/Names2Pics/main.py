import os

from PIL import Image
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook, load_workbook

SOURCE_EXCEL = 'names.xlsx'
OUTPUT_PATH = 'output/'
PICTURE_SIZE = (1000, 1000)
FONT_NAME = 'Hiragino Sans GB W6.ttc'
FONT_SIZE = 280
FONT_COLOR = (0, 0, 0)
BACKGROUND_COLOR = ()


def main():
    if not os.path.exists(OUTPUT_PATH):
        os.makedirs(OUTPUT_PATH)

    names = []

    workbook = load_workbook(SOURCE_EXCEL)
    worksheet = workbook.active

    names = []
    for row in range(worksheet.max_row):
        cell_name = 'A{}'.format(row+1)
        cell_value = worksheet[cell_name].value

        names.append(cell_value)

    font = ImageFont.truetype(FONT_NAME, FONT_SIZE)

    for i, name in enumerate(names):
        image = Image.new('RGB', PICTURE_SIZE, 'white')
        draw = ImageDraw.Draw(image)

        width, height = font.getsize(name)
        position = (PICTURE_SIZE[0]/2-width/2, PICTURE_SIZE[1]/2-height/2)

        draw.text(position, name, font=font, fill=FONT_COLOR)

        image.save(os.path.join(OUTPUT_PATH, '{}.jpg').format(i))



if __name__ == '__main__':
    main()
